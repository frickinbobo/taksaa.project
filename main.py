import sys
sys.dont_write_bytecode = True
import os
import uuid
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'settings')
import django
django.setup()
from db.models import *
from django.core import serializers
from flask import Flask, render_template, request, redirect, session, jsonify, send_from_directory
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import json
from datetime import datetime as dt
import dropbox


app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'media/'
app.secret_key = "f2fe5f5584862090a200f995e9fc79f6"
# DROPBOX_ACCESS_TOKEN = 'sl.u.AFcuxKd_LTTXUm4YoaPblml8YfKLLX7PcnTLjfE8cFBRueWT25lcG6sKTkvNteVRGzytIQFZaAdN-CLR-B3Cet6YFttWpMxNwSHSPvRTvhu1J5NvvRc_iTHWJdrTsqpGs9A48xaHyhwRzu5U0x_IpSWbbDuIlbHGCJVsHUCt6TncrZDqm34fZVU3QLULu4mkEqtco7YOU78BRBIxGkgBpNI4gTg4C9-yZiI_Tz6P-56W8ToigfVRbFVrlXc8izh6VoXtIBDck1PS1hFHdEdjNJ9pekCz-WzKN2WxT3wT4mFTWGzNkZIr-6HHGyBbxilmhAbEGqfhySN0pGSYGX4EVo5cfwNZIXfwPXJloE8mjRnaRxZYuxzc4rz_FaH8rgjhrDaux9gCLfNCsSYyOYMbJPo4aaUQn7I58zkKwMw3IHAzogICSfLBqgMPvtVOD2DG3SAW7hNrq7d426_jSyWO0NgcaHRpowRh785FeCC9OptfrTQuM9CQk0mna5XhwxZAuXBSxTQqQmPBDdp_1AiM0gkkOMUCYPQ9Ke0-thr3YFfJ3KlnKDbreoprio6y0DOQpZc4z4NkXy4Z3vC2pgNLmeXVzbtqGWxdjNlfH08QQBfAERhSpuhtwrEAeewmRdMV9yFPCKcRIuIf259Dm506riXq-nud6a7ggbd4fv4qFD2D7mcSw63rkTJD6lScOof0zy4FkzbmLrd6vcutjQylVEqcLo92CfjqimEEXz6XLoLTJmZf-dhAc1TN404Dmx4_XLYdf5caw8gJRGe_yEoOJPFqzTfTJ5b9Q8AZ6L4ckMdyr9cm3wmHTQHY9GWr__gAH0yWfXWtAX7I8TLaZJxBsOXR4P8Yb4XAibjzxNdgccvL_IQUpqfSNImiAoenhAhJE2MJvXYNzQVSFb15h0tjyQuY2iwaKVFn_vmRbP3U3TDVy7r63PidiZDQhC8mgE6hkccvfofLAInWYn65zEY2qI2BbtDVx6dKbL22jjvTLzGcNXSEOBa955Pcx_NxpfTC_KKE8ap120IMTtso12opd9MKCkShVkM3pjj5sT4TtHM0IEhGDqkTOU8WlLaS-OUaCfUiaaLPsuucIqX5XcJttMH7j2YY4TPUBdfQmDBVX9Wi3U_oSfHmiXcRatWk9iu7lh8HNn5e4da-DHdoxW_b_McZwnd-qCZjH2b0pX0HnJ5c7mBkW0cMPGSKdtlOcO9vXt7L6G1UreabJGYQN6XrfBtpeB3Fm_cmLFxDcoBTU0hymF2uBbF4cN6_rzJfsK-82JEuu3hKAKy7gGIneejG7yTzVakxyxbH1oAjWx5JnVWsXjlQtkk_dVsu8sa2RZYinCwxJRIhmmuLf6gMLkF1mkU1zC0qMpgTfrrYCWu9WwJOxchbufgSpheg9YdwqZLuwJNLOwvgvrvqZIjujE8TmRs1jOw0nBTGQKoVPNM1Xmmnew'
DROPBOX_ACCESS_TOKEN = "sl.u.AFeCkw8IoP1l71DfjLEpMuqF1ZNwrx94dqdVZOIPAFx6DKmzzqEbGsfcfbsRXwqNa1tExfZRuz5E5JLXDSbuD3Z4SjBGR8RSxtSlwlZMZaG46eZdpvI68d4cA-4a0VHKuDJORrTU1FM3GdpUB5zEPSNNuWYh_EiQ5FfxqDwtEGnONq2Xrf1Ouglc69JaqDMkXIDwgb-fFenvyW1qltfHuPONMXK75aOdxhEY_wUc2lL15_G-p-Ao-zOyCADz3peb05EdQwQUYd_XRwWSJC93ifW_FkEBYVP1Ez7xPN8jZvNpbX1VhO2pnHYFslwED9lIqQHOBou9uBSV6jq3XubMkQz-bRN6LTJYilIowrGRV2-q3D5ZXw7-OTLNmnOjnDH8g6S7hOvculmRfNHZL6zJWVdJEXSgajJrT5WmGN_OfqOTcVc61xiaUzVbAqiKnDoe-BEbKPGUC0PLMfthYKFoqfFODDABLrChTkM0pkgxEH2ZwJTY7pjhZPrY58TsidDYXAgH2vgnwKOMFSzcpoWAPce9V7WJHrtKWY2b5jY7cEALG3_t6zRVd59_IymxIKSs0SZh5f2ha6G2iDEtrBmdBp8oEvt6uOwAfg0K5R721Cjn56Vpdy5oB5tWsXB0vgEbiCCBFklZrBhcfcKTajALLIXqDCrF1AN2x8ww-GAootnVa0z5ONYJb1xcQXHZBE9qpPCghZTk2kiZpC0pmsuxpOvVHlJgn9rUPMXEDWvUFsA5HUzUbKvFX5RvwcVB6341yeeP9Yb848wqkHBRE2Bnz5DvDS-nbLVXO_aM0w1dyy2JhdLOb9l6fUHV3dfuPWm-GVvrRse_bDztjs9r1ILqZj6DVjZ31smOeRd3zEQv8duHFwTJP3NjIvaspjI8e9lhMY16czjR_GzFDM-TzKp68QoLAaPx7ISjGRdQnwkWQDdJgAAqoMoXxC6y_ZBWxsMFeXNCOhJayxKs6kC6yJADAiBOYqzaYiSjhUHlL14fEwYGkmQteD-tOUpqtYVtVMBju9Fl3aU5guxw3C3eYyz7ZRphrFw8rYHLw5Licys7yFtB7VNRjEZTXRGuK8YL80PdlGpklgRV6PD_af7HuxaE_qUgllu_rpR_cQc81wgzJPEjgjoyxzCJv-c_VES0-I-vCtkOBcYgxotgf9gmU7hvW4X-jbjA3DoC8t-DWEfCJsNF9kmk2pwruFSxhlDR3eyyaNeLmeWkYVxgP1cLNZqOXTH60iw_NneMxVny_08q7kDN79Y91CgC-aO2jwsszgcQ6XYoDGZ3ABLcGO8YrB0BUFI1khh4xe8uxRpkAUhf4-ovLLvdyeEdX-kz6Kx02py6Q-0grRZUKZ6J4E2b000rKg9T2-qBHhM0l7uwI6uF0iDYdgCnfwynBnhNACT-qIX7qqdFWViP2eYPPh0Ge3M8ys1EeyepPowDt9ESfJJjP5YxvA"
DROPBOX_REFRESH_TOKEN = "2C8moGTcI60AAAAAAAAAI46uVfdRf7w2zKNVh1727sY"
DROPBOX_APP_KEY = "hel6ah0l1cva71p"
DROPBOX_APP_SECRET = "rwogqdb9ck5ujmc"
# Index Route
@app.route('/', methods=['POST', 'GET'])
def index():
  return render_template('index.html')

# Admin Route
@app.route('/login/', methods=['POST', 'GET'])
def login():
  response = {}
  if request.method == 'POST':
    username = request.form['username']
    password = request.form['password']

    user = User.objects.filter(username=username).first()
    if user and check_password_hash(user.password, password):
      session['username'] = user.username
    else:
      response['type'] = 'error'
      response['message'] = 'Login failed! Please check your username or password and try again!'
  if session.get('username'):
    return redirect('/admin/band/')
    

  return render_template('admin/login.html', response=response)


@app.route('/logout/', methods=['POST', 'GET'])
def logout():
  session['username'] = None
  return redirect('/login/')

@app.route('/change-password/', methods=['POST', 'GET'])
def change_password():
  session['username'] = None
  response = {}
  if request.method == 'POST':
    username = request.form['username']
    old_password = request.form['old-password']
    new_password = request.form['new-password']
    confirm_new_password = request.form['confirm-new-password']
    user = User.objects.filter(username=username).first()
    if new_password == confirm_new_password:
      if user and check_password_hash(user.password, old_password):
        User.objects.filter(pk=user.id).update(password=generate_password_hash(new_password))
        return redirect('/login/')
      else:
        response['type'] = 'error'
        response['message'] = 'Password change failed! Please check your username or old password and try again!'
    else:
      response['type'] = 'error'
      response['message'] = 'Please repeat your new password!'

  return render_template('admin/change-password.html', response=response)


@app.route('/admin/', methods=['POST', 'GET'])
def dashboard():
  return redirect('/admin/band/')

@app.route('/admin/band/', methods=['POST', 'GET'])
def band():
  if not session.get('username'):
    return redirect('/login/')

  message = ''
  if request.method == 'POST':
    if request.form['button-form'] == 'add':
      name = request.form['name']
      description = "<br>".join(request.form['description'].replace('\r', '').split('\n'))
      spotify = request.form['spotify']
      instagram = request.form['instagram']
      if 'show' in request.form.keys():
        show = request.form['show']
      else:
        show = "off"
      cover = request.files['cover']
      cover.filename = secure_filename(uuid.uuid4().hex + '.' + cover.filename.split('.')[-1])
      real_path_cover = '/band/band-covers/'+cover.filename
      shared_link_cover = uploader(cover, '/band/band-covers/')
      logo = request.files['logo']
      logo.filename = secure_filename(uuid.uuid4().hex + '.' + logo.filename.split('.')[-1])
      real_path_logo = '/band/band-logos/'+logo.filename
      shared_link_logo = uploader(logo, '/band/band-logos/')
      # cover.filename = secure_filename(uuid.uuid4().hex + '.' + cover.filename.split('.')[-1])
      # real_path_cover = os.path.join(app.config['UPLOAD_FOLDER']+'band/band-covers/', cover.filename)
      # cover.save(real_path_cover)
      # logo = request.files['logo']
      # logo.filename = secure_filename(uuid.uuid4().hex + '.' + logo.filename.split('.')[-1])
      # real_path_logo = os.path.join(app.config['UPLOAD_FOLDER']+'band/band-logos/', logo.filename)
      # logo.save(real_path_logo)
      band = Band.objects.create(name=name, description=description, spotify=spotify, instagram=instagram, show=show, cover=shared_link_cover, cover_path=real_path_cover, logo=shared_link_logo, logo_path=real_path_logo)
      message = {'status': 'success', 'message': 'Band has been added successfully!'}
    elif request.form['button-form'] == 'edit':
      band = Band.objects.filter(pk=request.form['id']).first()
      id = request.form['id']
      name = request.form['name']
      description = "<br>".join(request.form['description'].replace('\r', '').split('\n'))
      spotify = request.form['spotify']
      instagram = request.form['instagram']
      if 'show' in request.form.keys():
        show = request.form['show']
      else:
        show = "off"
      cover = request.files['cover']
      logo = request.files['logo']
      
      if cover.filename and logo.filename:
        destroyer(band.cover_path)
        cover.filename = secure_filename(uuid.uuid4().hex + '.' + cover.filename.split('.')[-1])
        real_path_cover = '/band/band-covers/'+cover.filename
        shared_link_cover = uploader(cover, '/band/band-covers/')
        destroyer(band.logo_path)
        logo.filename = secure_filename(uuid.uuid4().hex + '.' + logo.filename.split('.')[-1])
        real_path_logo = '/band/band-logos/'+logo.filename
        shared_link_logo = uploader(logo, '/band/band-logos/')
        edited = Band.objects.filter(pk=id).update(name=name, description=description, spotify=spotify, instagram=instagram, show=show, cover=shared_link_cover, cover_path=real_path_cover, logo=shared_link_logo, logo_path=real_path_logo)
      elif cover.filename:
        destroyer(band.cover_path)
        cover.filename = secure_filename(uuid.uuid4().hex + '.' + cover.filename.split('.')[-1])
        real_path_cover = '/band/band-covers/'+cover.filename
        shared_link_cover = uploader(cover, '/band/band-covers/')
        edited = Band.objects.filter(pk=id).update(name=name, description=description, spotify=spotify, instagram=instagram, show=show, cover=shared_link_cover, cover_path=real_path_cover)
      elif logo.filename:
        destroyer(band.logo_path)
        logo.filename = secure_filename(uuid.uuid4().hex + '.' + logo.filename.split('.')[-1])
        real_path_logo = '/band/band-logos/'+logo.filename
        shared_link_logo = uploader(logo, '/band/band-logos/')
        edited = Band.objects.filter(pk=id).update(name=name, description=description, spotify=spotify, instagram=instagram, show=show, logo=shared_link_logo, logo_path=real_path_logo)
      else:
        edited = Band.objects.filter(pk=id).update(name=name, description=description, spotify=spotify, instagram=instagram, show=show)
      message = {'status': 'success', 'message': 'Band has been edited successfully!'}
    if request.form['button-form'] == 'delete':
      band = Band.objects.filter(pk=request.form['id'])
      if band:
        destroyer(band.first().cover_path)
        destroyer(band.first().logo_path)
        band.delete()
        message = {'status': 'success', 'message': 'Band has been deleted successfully!'}

  bands = Band.objects.all()
  return render_template('admin/band.html', bands=bands, message=message)


@app.route('/admin/band/<int:id>/', methods=['POST', 'GET'])
def item(id):
  if not session.get('username'):
    return redirect('/login/')
  message = {}
  response = {}
  band = Band.objects.filter(pk=id).first()
  if band:
    response['band'] = band
    
  if request.method == 'POST':
    if request.form['button-form'] == 'add':
      name = request.form['name']
      description = "<br>".join(request.form['description'].replace('\r', '').split('\n'))
      if 'use-stock' in request.form.keys():
        use_stock = request.form['use-stock']
      else:
        use_stock = "off"
      if 'show' in request.form.keys():
        show = request.form['show']
      else:
        show = "off"
      new_item = Item.objects.create(band=band, name=name, description=description, use_stock=use_stock, show=show)
      sizes = []
      prices = []
      stocks = []
      for size in request.form.keys():
        d = {}
        if 'size-' in size:
          sizes.append(request.form[size])
        if 'price-' in size:
          prices.append(request.form[size])
        if 'stock-' in size:
          stocks.append(request.form[size])
        
      for i in range(len(sizes)):
        new_size = ItemSize.objects.create(item=new_item, price=prices[i], size=sizes[i], stock=stocks[i])

      # images = request.files.getlist('images')
      # dir = app.config['UPLOAD_FOLDER']+'item-images/'+ str(new_item.id) + '/'
      # if not os.path.exists(dir):
      #     os.makedirs(dir)

      # for image in images:
      #   image.filename = secure_filename(uuid.uuid4().hex + '.' + image.filename.split('.')[-1])
      #   real_path = os.path.join(dir, image.filename)
      #   image.save(real_path)
      #   ItemImage.objects.create(item=new_item, image=real_path)
      images = request.files.getlist('images')

      for image in images:
        image.filename = secure_filename(uuid.uuid4().hex + '.' + image.filename.split('.')[-1])
        real_path = '/item-images/'+ str(new_item.id)+'/'+image.filename
        shared_link = uploader(image, '/item-images/'+ str(new_item.id)+'/')
        ItemImage.objects.create(item=new_item, image=shared_link, path=real_path)

      message = {'status': 'success', 'message': 'Clothes has been added successfully!'}

    elif request.form['button-form'] == 'edit':
      item = Item.objects.filter(pk=request.form['id'])
      if len(item) > 0:
        name = request.form['name']
        description = "<br>".join(request.form['description'].replace('\r', '').split('\n'))
        if 'use-stock' in request.form.keys():
          orders = Order.objects.all()
          for order in orders:
            if order.status == 'In Progress':
              for orderitem in order.orderitem_set.all():
                if orderitem.item_size.item.id == item.first().id:
                  before_stock = ItemSize.objects.filter(pk=orderitem.item_size.id).first().stock
                  ItemSize.objects.filter(pk=orderitem.item_size.id).update(stock=before_stock-orderitem.quantity)
          use_stock = request.form['use-stock']
        else:
          orders = Order.objects.all()
          for order in orders:
            if order.status == 'In Progress':
              for orderitem in order.orderitem_set.all():
                if orderitem.item_size.item.id == item.first().id:
                  before_stock = ItemSize.objects.filter(pk=orderitem.item_size.id).first().stock
                  ItemSize.objects.filter(pk=orderitem.item_size.id).update(stock=before_stock+orderitem.quantity)
          use_stock = "off"
        
        if 'show' in request.form.keys():
          show = request.form['show']
        else:
          show = "off"
        print(show)
        item.update(name=name, description=description, use_stock=use_stock, show=show)

      message = {'status': 'success', 'message': 'Clothes has been edited successfully!'}
      
    elif request.form['button-form'] == 'delete':
      item = Item.objects.filter(pk=request.form['id'])
      if item:
        images = item.first().itemimage_set.all()
        for image in images:
          destroyer(image.path)
        destroyer('/item-images/'+str(item.first().id))
        item.delete()

      message = {'status': 'success', 'message': 'Clothes has been deleted successfully!'}
    
    elif request.form['button-form'] == 'add-image':
      item = Item.objects.filter(pk=request.form['id']).first()
      if item:
        images = request.files.getlist('image-add')
        for image in images:
          image.filename = secure_filename(uuid.uuid4().hex + '.' + image.filename.split('.')[-1])
          real_path = '/item-images/'+ str(item.id)+'/'+image.filename
          shared_link = uploader(image, '/item-images/'+ str(item.id)+'/')
          ItemImage.objects.create(item=item, image=shared_link, path=real_path)

      message = {'status': 'success', 'message': 'Clothes image(s) has been added successfully!'}

    elif request.form['button-form'] == 'edit-image':
      if request.form['id']:
        item_image = ItemImage.objects.filter(pk=request.form['id']).first()
        if request.files['image-edit']:
          image = request.files['image-edit']
          destroyer(item_image.path)
          image.filename = secure_filename(uuid.uuid4().hex + '.' + image.filename.split('.')[-1])
          real_path = '/item-images/'+ str(item_image.id)+'/'+image.filename
          shared_link = uploader(image, '/item-images/'+ str(item_image.id)+'/')
          ItemImage.objects.filter(pk=request.form['id']).update(image=shared_link, path=real_path)

        message = {'status': 'success', 'message': 'Clothes image(s) has been edited successfully!'}


  if band:
    items = band.item_set.all()
    response['items'] = [{'item': item, 'size': item.itemsize_set.all(), 'image': item.itemimage_set.all()} for item in items]

  return render_template('admin/item.html', response=response, message=message)

@app.route('/admin/order/', methods=['POST', 'GET'])
def order():
  if not session.get('username'):
    return redirect('/login/')

  message = {}
  if request.method == 'POST':
    if request.form['button-form'] == 'add':
      name = request.form['name']
      phone = request.form['phone']
      email = request.form['email']
      address = request.form['address']
      type = request.form['type']
      status = request.form['status']
      receipt = request.files['receipt']
      # receipt.filename = secure_filename(uuid.uuid4().hex + '.' + receipt.filename.split('.')[-1])
      # real_path_receipt = os.path.join(app.config['UPLOAD_FOLDER']+'customer-receipt/', receipt.filename)
      # receipt.save(real_path_receipt)
      receipt.filename = secure_filename(uuid.uuid4().hex + '.' + receipt.filename.split('.')[-1])
      real_path = '/customer-receipt/'+receipt.filename
      shared_link = uploader(receipt, '/customer-receipt/')
      new_order = Order.objects.create(customer_name=name, customer_phone=phone, customer_email=email, customer_address=address, status=status, type=type, customer_receipt=shared_link, customer_receipt_path=real_path)
      items = []
      quantities = []
      for size in request.form.keys():
        d = {}
        if 'id-' in size:
          items.append(ItemSize.objects.filter(pk=request.form[size]).first())
        if 'quantity-' in size:
          quantities.append(request.form[size])
        
      for i in range(len(items)):
        item_size = ItemSize.objects.filter(pk=items[i].id).first()
        before_stock = item_size.stock
        if item_size.item.use_stock == 'on':
          ItemSize.objects.filter(pk=item_size.id).update(stock=before_stock-int(quantities[i]))
        new_size = OrderItem.objects.create(order=new_order, item_size=items[i], quantity=quantities[i])

      message = {'status': 'success', 'message': 'Order has been added successfully!'}

    if request.form['button-form'] == 'edit':
      order_id = request.form['id']
      order = Order.objects.filter(pk=order_id).first()
      if order:
        name = request.form['name']
        phone = request.form['phone']
        email = request.form['email']
        address = request.form['address']
        type = request.form['type']
        status = request.form['status']
        receipt = request.files['receipt']
        saved_clothes = [order_item.item_size.id for order_item in OrderItem.objects.filter(order=order)]
        items = []
        quantities = []
        for size in request.form.keys():
          d = {}
          if 'id-' in size:
            items.append(ItemSize.objects.filter(pk=request.form[size]).first())
          if 'quantity-' in size:
            quantities.append(request.form[size])

        new_clothes = [clothe.id for clothe in items]

        for i in range(len(items)):
          if not items[i].id in saved_clothes:
            item_size = ItemSize.objects.filter(pk=items[i].id).first()
            before_stock = item_size.stock
            if item_size.item.use_stock == 'on':
              ItemSize.objects.filter(pk=item_size.id).update(stock=before_stock-int(quantities[i]))
            OrderItem.objects.create(order=order, item_size=items[i], quantity=quantities[i])
            saved_clothes.append(items[i].id)
          else:
            orderitem = OrderItem.objects.filter(item_size=items[i])
            item_size = ItemSize.objects.filter(pk=items[i].id).first()
            before_stock = items[i].stock
            if item_size.item.use_stock == 'on':
              ItemSize.objects.filter(pk=item_size.id).update(stock=before_stock-(int(quantities[i])-orderitem.first().quantity))
            orderitem.update(quantity=quantities[i])

        for id in saved_clothes:
          item_size = ItemSize.objects.filter(pk=id).first()
          if not item_size.id in new_clothes:
            if status == 'In Progress':
              before_stock = item_size.stock
              if item_size.item.use_stock == 'on':
                ItemSize.objects.filter(pk=item_size.id).update(stock=before_stock+int(quantities[i]))
            OrderItem.objects.filter(item_size=item_size).delete()


        if receipt.filename:
          destroyer(receipt.customer_receipt_path)
          receipt.filename = secure_filename(uuid.uuid4().hex + '.' + receipt.filename.split('.')[-1])
          real_path = '/customer-receipt/'+receipt.filename
          shared_link = uploader(receipt, '/customer-receipt/')
          edited = Order.objects.filter(pk=order_id).update(customer_name=name, customer_phone=phone, customer_email=email, customer_address=address, status=status, type=type, customer_receipt=shared_link, customer_receipt_path=real_path)
        else:
          edited = Order.objects.filter(pk=order_id).update(customer_name=name, customer_phone=phone, customer_email=email, customer_address=address, status=status, type=type)
          print(Order.objects.filter(pk=order_id))
        message = {'status': 'success', 'message': 'Order has been edited successfully!'}

    if request.form['button-form'] == 'delete':
      order = Order.objects.filter(pk=request.form['id'])
      if order:
        if order.first().status == 'In Progress':
          for orderitem in order.first().orderitem_set.all():
            item_size = ItemSize.objects.filter(pk=orderitem.item_size.id)
            before_stock = item_size.first().stock
            item_size.update(stock=before_stock+orderitem.quantity)
        try:
          destroyer(order.customer_receipt_path)
        except:
          pass
        order.delete()
        message = {'status': 'success', 'message': 'Order has been deleted successfully!'}

  orders = []
  for order in Order.objects.all():
    single_order = {}
    single_order['order'] = order
    single_order['order_item'] = order.orderitem_set.all()
    orders.append(single_order)
  
  clothes = []
  for item in Item.objects.all():
    clothe = {}
    clothe['item'] = item
    clothe['images'] = item.itemimage_set.all()
    clothe['sizes'] = item.itemsize_set.all()
    clothes.append(clothe)

  return render_template('admin/order.html', orders=orders, clothes=clothes, message=message)


# API
# Get
@app.route('/api/get/bands/', methods=['GET'])
def api_get_bands():
  bands = serializers.serialize('json', Band.objects.all())
  return jsonify(bands)

@app.route('/api/get/orders/', methods=['GET'])
def api_get_orders():
  orders = []
  for order in Order.objects.all():
    single_order = {}
    single_order['order'] = serializers.serialize('json', [order])
    single_order['order_item'] = serializers.serialize('json', order.orderitem_set.all())
    orders.append(single_order)
  
  clothes = []
  for item in Item.objects.all():
    clothe = {}
    clothe['item'] = serializers.serialize('json', [item])
    clothe['images'] = serializers.serialize('json', item.itemimage_set.all())
    clothe['sizes'] = serializers.serialize('json', item.itemsize_set.all())
    clothes.append(clothe)
  
  return jsonify({'orders': orders, 'clothes': clothes})

@app.route('/api/get/band/<int:id_band>/items/', methods=['GET'])
def api_get_band_items(id_band):
  band = Band.objects.filter(pk=id_band).first()
  if band:
    items = band.item_set.all()
    data = [{'item': serializers.serialize('json', [item]), 'size': serializers.serialize('json', item.itemsize_set.all()), 'image': serializers.serialize('json', item.itemimage_set.all())} for item in items]
    return jsonify(data)

@app.route('/api/get/item-data/size/<int:id_item>/', methods=['GET'])
def api_get_itemsize(id_item):
  item = serializers.serialize('json', Item.objects.filter(pk=id_item))
  sizes = serializers.serialize('json', Item.objects.filter(pk=id_item).first().itemsize_set.all())
  data = {'item': item, 'sizes': sizes}
  return jsonify(data)


@app.route('/api/get/item-data/image/<int:id_item>/', methods=['GET'])
def api_get_itemimage(id_item):
  item = serializers.serialize('json', Item.objects.filter(pk=id_item))
  images = serializers.serialize('json', Item.objects.filter(pk=id_item).first().itemimage_set.all())
  data = {'item': item, 'images': images}
  return jsonify(data)

@app.route('/api/get/order/<int:id_order>/items/', methods=['GET'])
def api_get_orderitem(id_order):
  order = serializers.serialize('json', Order.objects.filter(pk=id_order))
  # items = serializers.serialize('json', Order.objects.filter(pk=id_order).first().orderitem_set.all())
  clothes = []
  for clothe in Order.objects.filter(pk=id_order).first().orderitem_set.all():
    clothes.append({'clothe': serializers.serialize('json', ItemSize.objects.filter(pk=clothe.item_size.id)), 'quantity': clothe.quantity, 'name': clothe.item_size.item.name, 'image': clothe.item_size.item.itemimage_set.all().first().image})
  
  data = {'order': order, 'clothes': clothes}
  return jsonify(data)

@app.route('/api/get/all-band/', methods=['GET'])
def api_get_allband():
  bands = serializers.serialize('json', Band.objects.filter(show='on'))
  return jsonify(bands)

@app.route('/api/get/all-clothe/', methods=['GET'])
def api_get_all_clothes():
  data=[]
  clothes = Item.objects.filter(show='on')
  for clothe in clothes:
    temp = {}
    temp['clothe'] = serializers.serialize('json', [clothe])
    temp['sizes'] = serializers.serialize('json', ItemSize.objects.filter(item=clothe))
    temp['images'] = serializers.serialize('json', ItemImage.objects.filter(item=clothe))
    data.append(temp)
  
  return jsonify(data)

@app.route('/api/get/band/<int:id_band>/clothes/', methods=['GET'])
def api_get_band_clothes(id_band):
  data = {}
  band = Band.objects.filter(pk=id_band)
  if band:
    data['band'] = serializers.serialize('json', band)
    clothes = Item.objects.filter(band=band.first(), show='on')
    data['clothes'] = []
    for clothe in clothes:
      temp = {}
      temp['clothe'] = serializers.serialize('json', [clothe])
      temp['sizes'] = serializers.serialize('json', ItemSize.objects.filter(item=clothe))
      temp['images'] = serializers.serialize('json', ItemImage.objects.filter(item=clothe))
      data['clothes'].append(temp)
      
  else:
    data = {'status': 304}
  return jsonify(data)

@app.route('/api/get/clothes/<int:id_clothe>/', methods=['GET'])
def api_get_clothes(id_clothe):
  data = {}
  clothe = Item.objects.filter(pk=id_clothe, show='on')
  if clothe:
    data['clothe'] = serializers.serialize('json', clothe)
    data['sizes'] = serializers.serialize('json', clothe.first().itemsize_set.all())
    data['images'] = serializers.serialize('json', clothe.first().itemimage_set.all())

  return jsonify(data)
# Delete
@app.route('/api/delete/item-data/size/<int:id_item>/', methods=['DELETE'])
def delete_item(id_item):
  if session.get('username'):
    if request.method == 'DELETE':
      item = ItemSize.objects.filter(pk=id_item).first().delete()
  return jsonify({})

@app.route('/api/delete/item-data/image/<int:id_image>/', methods=['DELETE'])
def delete_image(id_image):
  if session.get('username'):
    if request.method == 'DELETE':
      item = ItemImage.objects.filter(pk=id_image).first()
      destroyer(item.path)
      item.delete()
  return jsonify({})

# Put
@app.route('/api/put/item-data/size/<int:id_item>/', methods=['PUT'])
def put_item(id_item):
  if session.get('username'):
    if request.method == 'PUT':
      size = request.get_json()['size']
      price = request.get_json()['price']
      stock = request.get_json()['stock']
      item = ItemSize.objects.filter(pk=id_item).update(size=size, price=price, stock=stock)
  
  return jsonify({})

# Post
@app.route('/api/post/item-data/size/<int:id_item>/', methods=['POST'])
def post_item(id_item):
  if session.get('username'):
    if request.method == 'POST':
      size = request.get_json()['size']
      price = request.get_json()['price']
      stock = request.get_json()['stock']
      item = ItemSize.objects.create(item=Item.objects.filter(pk=id_item).first(), size=size, price=price, stock=stock)
  
  return jsonify({})

@app.route('/api/post/order/', methods=['POST'])
def post_order():
  if request.method == 'POST':
    try:
      name = request.form['name']
      address = request.form['address']
      phone = request.form['phone']
      email = request.form['email']
      type = request.form['type']
      receipt = request.files['receipt']
      clothes = json.loads(request.form['orders'])
      receipt.filename = secure_filename(uuid.uuid4().hex + '.' + receipt.filename.split('.')[-1])
      real_path = '/customer-receipt/'+receipt.filename
      shared_link = uploader(receipt, '/customer-receipt/')
      
      order = Order.objects.create(customer_name=name, customer_phone=phone, customer_email=email, customer_address=address, status='In Progress', type=type, customer_receipt=shared_link, customer_receipt_path=real_path)
      for key in clothes.keys():
        item_size = ItemSize.objects.filter(pk=key).first()
        order_item = OrderItem.objects.create(order=order, item_size=item_size, quantity=clothes[key]['quantity'])
      return jsonify({'status': 200, 'message': 'Success!'})
    except Exception as e:
      return jsonify({'status': 500, 'message': e})

# Expose Media Folder
@app.get('/media/<path:path>')
def send_media(path):
    return send_from_directory(
        directory=app.config['UPLOAD_FOLDER'], path=path
    )


@app.route('/admin/convert-to-excel/order/')
def convert_to_excel():
  import openpyxl
  from openpyxl.styles import Alignment
  from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
  from openpyxl.utils import get_column_letter
  

  wb = openpyxl.Workbook()

  sheet = wb.active


  sheet.merge_cells('A1:A2')
  sheet['A1'] = "No"
  sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

  sheet.merge_cells('B1:B2')
  sheet['B1'] = "Name"
  sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')

  sheet.merge_cells('C1:C2')
  sheet['C1'] = "Phone"
  sheet['C1'].alignment = Alignment(horizontal='center', vertical='center')

  sheet.merge_cells('D1:D2')
  sheet['D1'] = "Email"
  sheet['D1'].alignment = Alignment(horizontal='center', vertical='center')

  sheet.merge_cells('E1:E2')
  sheet['E1'] = "Address"
  sheet['E1'].alignment = Alignment(horizontal='center', vertical='center')

  sheet.merge_cells('F1:F2')
  sheet['F1'] = "Order Type"
  sheet['F1'].alignment = Alignment(horizontal='center', vertical='center')

  sheet.merge_cells('G1:L1')
  sheet['G1'] = 'Clothes Ordered'
  sheet['G1'].alignment = Alignment(horizontal='center', vertical='center')

  sheet['G2'] = 'No'
  sheet['G2'].alignment = Alignment(horizontal='center', vertical='center')
  sheet['H2'] = 'Id'
  sheet['H2'].alignment = Alignment(horizontal='center', vertical='center')
  sheet['I2'] = 'Clothe'
  sheet['I2'].alignment = Alignment(horizontal='center', vertical='center')
  sheet['J2'] = 'Size'
  sheet['J2'].alignment = Alignment(horizontal='center', vertical='center')
  sheet['K2'] = 'Quantity'
  sheet['K2'].alignment = Alignment(horizontal='center', vertical='center')
  sheet['L2'] = 'Price/pcs'
  sheet['L2'].alignment = Alignment(horizontal='center', vertical='center')

  sheet.merge_cells('M1:M2')
  sheet['M1'] = 'Total'
  sheet['M1'].alignment = Alignment(horizontal='center', vertical='center')

  sheet.merge_cells('N1:N2')
  sheet['N1'] = 'Status'
  sheet['N1'].alignment = Alignment(horizontal='center', vertical='center')

  sheet.merge_cells('O1:O2')
  sheet['O1'] = 'Receipt'
  sheet['O1'].alignment = Alignment(horizontal='center', vertical='center')

  start = 3
  last_cell = ''
  orders = Order.objects.all()
  for i, order in enumerate(orders):
    order_items = order.orderitem_set.all()
    merged_by = len(order_items)
    sheet.merge_cells('A'+str(start)+':A'+str(start+merged_by-1))
    sheet['A'+str(start)].alignment = Alignment(horizontal='center', vertical='center')
    sheet['A'+str(start)] = i + 1

    sheet.merge_cells('B'+str(start)+':B'+str(start+merged_by-1))
    sheet['B'+str(start)].alignment = Alignment(horizontal='left', vertical='center')
    sheet['B'+str(start)] = order.customer_name

    sheet.merge_cells('C'+str(start)+':C'+str(start+merged_by-1))
    sheet['C'+str(start)].alignment = Alignment(horizontal='left', vertical='center')
    sheet['C'+str(start)] = order.customer_phone
    
    sheet.merge_cells('D'+str(start)+':D'+str(start+merged_by-1))
    sheet['D'+str(start)].alignment = Alignment(horizontal='left', vertical='center')
    sheet['D'+str(start)] = order.customer_email

    sheet.merge_cells('E'+str(start)+':E'+str(start+merged_by-1))
    sheet['E'+str(start)].alignment = Alignment(horizontal='left', vertical='center')
    sheet['E'+str(start)] = order.customer_address
    
    sheet.merge_cells('F'+str(start)+':F'+str(start+merged_by-1))
    sheet['F'+str(start)].alignment = Alignment(horizontal='left', vertical='center')
    sheet['F'+str(start)] = order.type
    sum = 0
    for j, item in enumerate(order_items):
      sheet['G'+str(start + j)] = j + 1
      sheet['G'+str(start + j)].alignment = Alignment(horizontal='center', vertical='center')

      sheet['H'+str(start + j)] = item.item_size.item.id
      sheet['H'+str(start + j)].alignment = Alignment(horizontal='center', vertical='center')

      sheet['I'+str(start + j)] = item.item_size.item.name
      sheet['I'+str(start + j)].alignment = Alignment(horizontal='left', vertical='center')
      sheet['J'+str(start + j)] = item.item_size.size
      sheet['J'+str(start + j)].alignment = Alignment(horizontal='center', vertical='center')
      sheet['K'+str(start + j)] = item.quantity
      sheet['K'+str(start + j)].alignment = Alignment(horizontal='center', vertical='center')
      sheet['L'+str(start + j)] = item.item_size.price
      sheet['L'+str(start + j)].alignment = Alignment(horizontal='right', vertical='center')
      sum += (item.item_size.price * item.quantity)
    
    sheet.merge_cells('M'+str(start)+':M'+str(start+merged_by-1))
    sheet['M'+str(start)].alignment = Alignment(horizontal='right', vertical='center')
    sheet['M'+str(start)] = sum
    
    sheet.merge_cells('N'+str(start)+':N'+str(start+merged_by-1))
    sheet['N'+str(start)].alignment = Alignment(horizontal='center', vertical='center')
    sheet['N'+str(start)] = order.status

    sheet.merge_cells('O'+str(start)+':O'+str(start+merged_by-1))
    sheet['O'+str(start)].alignment = Alignment(horizontal='left', vertical='center')
    sheet['O'+str(start)] = order.customer_receipt

    if merged_by == 0:
      start += 1
    else:
      start += merged_by

  sheet['A'+str(start+2)] = 'Generated at '+dt.now().strftime("%A, %d %B %Y, %H:%M:%S")

  column_widths = []
  for row in sheet.iter_rows():
    for i, cell in enumerate(row):
      if i > 1:
        try:
          column_widths[i] = max(column_widths[i], len(str(cell.value)) + 4)
        except IndexError:
          column_widths.append(len(str(cell.value)) + 4)

  for i, column_width in enumerate(column_widths):
    sheet.column_dimensions[get_column_letter(i + 1)].width = column_width

  dir = app.config['UPLOAD_FOLDER']+'exports/'
  if not os.path.exists(dir):
      os.makedirs(dir)

  wb.save(app.config['UPLOAD_FOLDER']+'exports/'+dt.now().strftime("%A, %d %B %Y, %H:%M:%S")+".xlsx")

  return redirect('/'+app.config['UPLOAD_FOLDER']+'exports/'+dt.now().strftime("%A, %d %B %Y, %H:%M:%S")+".xlsx")


@app.route('/test/', methods=['POST', 'GET'])
def test():
  # user = User.objects.create(name='test', username='admin', password=generate_password_hash('admin'), role='test', status='test')
  import re
  if request.method == 'POST':
    data = "<br/>".join(request.form['texarea'].replace('\r', '').split('\n'))
    print(data)

  return render_template('admin/test.html')


def uploader(file, dest='/'):
  # Jp71E1dKgqIAAAAAAAAAATYjWWNUFw1UHUL2bW2kmvQRw5q3VQTdrZjLAZYcLaFQ
  dbx = dropbox.Dropbox(app_key=DROPBOX_APP_KEY, app_secret=DROPBOX_APP_SECRET, oauth2_refresh_token="Jp71E1dKgqIAAAAAAAAAATYjWWNUFw1UHUL2bW2kmvQRw5q3VQTdrZjLAZYcLaFQ")
  uploaded = dbx.files_upload(file.read(), dest+file.filename, mode=dropbox.files.WriteMode('overwrite'))
  shared_links = dbx.sharing_list_shared_links(uploaded.path_display)
  if len(shared_links.links)>0:
    return shared_links.links[0].url[:-1] + '1'
  else:
    shared_link = dbx.sharing_create_shared_link_with_settings(uploaded.path_display)
    return shared_link.url[:-1] + '1'

def destroyer(path):
  dbx = dropbox.Dropbox(app_key=DROPBOX_APP_KEY, app_secret=DROPBOX_APP_SECRET, oauth2_refresh_token="Jp71E1dKgqIAAAAAAAAAATYjWWNUFw1UHUL2bW2kmvQRw5q3VQTdrZjLAZYcLaFQ")
  dbx.files_delete_v2(path)

if __name__ == '__main__':
  app.run(debug=True, host='0.0.0.0')
  # app.run()